import pandas as pd
from rapidfuzz import fuzz, process
from openpyxl import load_workbook
import os

# === State Name to Abbreviation Mapping ===
us_state_to_abbrev = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR", "California": "CA",
    "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE", "Florida": "FL", "Georgia": "GA",
    "Hawaii": "HI", "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA",
    "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME", "Maryland": "MD",
    "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS",
    "Missouri": "MO", "Montana": "MT", "Nebraska": "NE", "Nevada": "NV",
    "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM", "New York": "NY",
    "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK",
    "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI", "South Carolina": "SC",
    "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX", "Utah": "UT", "Vermont": "VT",
    "Virginia": "VA", "Washington": "WA", "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
    "District of Columbia": "DC"
}

# === File paths ===
input_directory = 'input/'
output_directory = 'output/events/'
input_file_name = 'matched_trifind_usat_events_2025_v1.xlsx'
output_file_name = 'matched_trifind_usat_events_2025_v1_updated.xlsx'
input_path = os.path.join(output_directory, input_file_name)
output_path = os.path.join(output_directory, output_file_name)
os.makedirs(output_directory, exist_ok=True)

# === Load Input Files ===
trifind = pd.read_excel(os.path.join(input_directory, 'trifind_advanced_search_2025_results.xlsx'), sheet_name='All Events')
usat = pd.read_csv(os.path.join(input_directory, 'results_2025-06-20_12-00-59_python_event_data_offset_0_batch_1.csv'))
print("USAT Columns:", usat.columns.tolist())

# === Clean + Normalize ===
trifind = trifind[['title', 'city', 'state', 'location', 'race_type', 'usat_sanctioned', 'url', 'date']].dropna().rename(columns={
    'url': 'trifind_url',
    'usat_sanctioned': 'trifind_usat_sanctioned_flag'
})

# Convert full state names to abbreviations, default to 'Other'
trifind['state_abbrev'] = trifind['state'].apply(lambda s: us_state_to_abbrev.get(s, 'Other'))

# Log unmatched states
other_states = trifind[trifind['state_abbrev'] == 'Other']['state'].unique()
if len(other_states) > 0:
    print("⚠️ The following Trifind states were not recognized and categorized as 'Other':")
    for s in other_states:
        print(f" - {s}")

# Clean USAT data
usat = usat[['Name', '2LetterCode', 'RaceDate', 'ApplicationID', 'Status', 'RegistrationWebsite']].dropna()
usat = usat.rename(columns={'2LetterCode': 'usat_state'})
usat['parsed_date'] = pd.to_datetime(usat['RaceDate'], errors='coerce')
usat['usat_month'] = usat['parsed_date'].dt.month
usat['usat_year'] = usat['parsed_date'].dt.year
usat = usat[usat['parsed_date'].dt.year == 2025]

# Normalize titles
trifind['title_norm'] = trifind['title'].str.lower().str.strip()
usat['Name_norm'] = usat['Name'].str.lower().str.strip()

# Parse Trifind dates
trifind['parsed_date'] = pd.to_datetime(trifind['date'], errors='coerce')
trifind['month'] = trifind['parsed_date'].dt.month
trifind['year'] = trifind['parsed_date'].dt.year

# === Fuzzy Match Logic ===
all_matches = []
for _, tf in trifind.iterrows():
    if tf['state_abbrev'] == 'Other':
        continue

    usat_candidates = usat[usat['usat_state'] == tf['state_abbrev']]
    if not usat_candidates.empty:
        usat_match_norm, score_usat, _ = process.extractOne(tf['title_norm'], usat_candidates['Name_norm'], scorer=fuzz.ratio)
        usat_row = usat_candidates[usat_candidates['Name_norm'] == usat_match_norm].iloc[0]
        usat_name = usat_row['Name']
        usat_status = usat_row['Status']
        usat_state = usat_row['usat_state']
        usat_month = usat_row['usat_month']
        usat_year = usat_row['usat_year']
        usat_date = usat_row['RaceDate']
        usat_app_id = usat_row['ApplicationID']
        usat_reg_url = usat_row['RegistrationWebsite']
    else:
        usat_name = usat_state = usat_date = usat_app_id = usat_reg_url = usat_month = usat_year = None
        score_usat = 0

    trifind_sanctioned = str(tf['trifind_usat_sanctioned_flag']).strip().lower() == 'yes'
    match_score_high = score_usat > 90
    is_sanctioned = trifind_sanctioned or match_score_high
    sanction_discrepancy_flag = trifind_sanctioned != match_score_high

    if trifind_sanctioned and match_score_high:
        reason = "both"
    elif trifind_sanctioned and not match_score_high:
        reason = "flag_only"
    elif not trifind_sanctioned and match_score_high:
        reason = "score_only"
    else:
        reason = "neither"

    all_matches.append({
        'trifind_title': tf['title'],
        'trifind_city': tf['city'],
        'trifind_state': tf['state'],
        'trifind_location': tf['location'],
        'trifind_race_type': tf['race_type'],
        'trifind_date': tf['parsed_date'],
        'trifind_month': tf['month'],
        'trifind_year': tf['year'],
        'trifind_url': tf['trifind_url'],
        'trifind_usat_sanctioned_flag': tf['trifind_usat_sanctioned_flag'],

        'usat_name': usat_name,
        'usat_status': usat_status,
        'usat_state': usat_state,
        'usat_date': usat_date,
        'usat_month': usat_month,
        'usat_year': usat_year,
        'match_score_usat': score_usat,
        'matched_usat': match_score_high,
        'ApplicationID': usat_app_id,
        'RegistrationWebsite': usat_reg_url,
        'inferred_usat_sanctioned': is_sanctioned,
        'sanction_discrepancy_flag': sanction_discrepancy_flag,
        'reason_for_sanction': reason
    })

df_matches = pd.DataFrame(all_matches)

# === Score Bins Summary ===
bins = [0, 69, 79, 89, 94, 100]
labels = ['0–69', '70–79', '80–89', '90–94', '95–100']
df_matches['score_bin_usat'] = pd.cut(df_matches['match_score_usat'], bins=bins, labels=labels, include_lowest=True)
summary_usat = df_matches['score_bin_usat'].value_counts().sort_index().reset_index()
summary_usat.columns = ['match_score_bin_usat', 'count']
total_matches = len(df_matches)

# === Write to Existing Excel Workbook ===
wb = load_workbook(input_path)

if "match_data" in wb.sheetnames:
    del wb["match_data"]
ws = wb.create_sheet("match_data")

for col_idx, col_name in enumerate(df_matches.columns, 1):
    ws.cell(row=1, column=col_idx).value = col_name
for row_idx, row in enumerate(df_matches.itertuples(index=False), 2):
    for col_idx, value in enumerate(row, 1):
        ws.cell(row=row_idx, column=col_idx).value = value

# === Write Instructions Tab ===
if "Instructions" in wb.sheetnames:
    del wb["Instructions"]
ws_instr = wb.create_sheet("Instructions")

summary_usat_lines = ["📊 USAT Match Score Bin Summary:"]
for _, row in summary_usat.iterrows():
    summary_usat_lines.append(f"{row['match_score_bin_usat']}: {row['count']} matches")
summary_usat_lines.append(f"TOTAL: {total_matches} matches")

summary_flag = df_matches['trifind_usat_sanctioned_flag'].value_counts().to_string(index=True, header=False).split('\n')
summary_inferred = df_matches['inferred_usat_sanctioned'].value_counts().to_string(index=True, header=False).split('\n')
summary_discrepancy = df_matches['sanction_discrepancy_flag'].value_counts().to_string(index=True, header=False).split('\n')
summary_reason = df_matches['reason_for_sanction'].value_counts().to_string(index=True, header=False).split('\n')

instructions = [
    "⚠️ How to Refresh Pivot Tables After Data Update",
    "",
    "1. Go to the 'pivot_by_state' or 'pivot_by_event' sheet.",
    "2. Click anywhere inside the pivot table.",
    "3. In the Excel ribbon, click 'Data' > 'Refresh All'.",
    "",
    "📘 Column Descriptions in 'match_data':",
    "",
    "trifind_title               – Title of the event as listed on Trifind.",
    "trifind_url                 – URL to the event's Trifind page.",
    "trifind_date                – Parsed event date from Trifind.",
    "trifind_month/year          – Parsed month and year of the event.",
    "trifind_state               – Full state name from Trifind.",
    "trifind_usat_sanctioned_flag – 'Yes' if Trifind lists it as USAT sanctioned.",
    "usat_name                   – Closest matching event name from USAT.",
    "usat_state                  – USAT 2-letter state abbreviation.",
    "usat_date                   – Date of the matched USAT event.",
    "usat_month/year             – Parsed month/year of USAT event.",
    "match_score_usat            – Fuzzy match score (0–100) between Trifind and USAT titles.",
    "matched_usat                – True if match score is > 90.",
    "ApplicationID               – Unique USAT event ID (if matched).",
    "RegistrationWebsite         – USAT registration URL for the matched event.",
    "inferred_usat_sanctioned    – True if either Trifind says 'yes' or score > 90.",
    "sanction_discrepancy_flag   – True if Trifind and USAT disagree.",
    "reason_for_sanction         – Logic:",
    "    'both'       – Trifind = Yes AND Score > 90",
    "    'flag_only' – Trifind = Yes, Score ≤ 90",
    "    'score_only'– Trifind != Yes, Score > 90",
    "    'neither'    – Neither source indicates sanctioning",
    "",
    "📊 Summary Stats from This File:"
] + summary_usat_lines + [""] + [
    "🟢 Trifind USAT Sanctioned Flag (Raw):"] + summary_flag + [""] + [
    "🟢 Inferred USAT Sanctioned (Flag OR Score > 90):"] + summary_inferred + [""] + [
    "⚠️ Sanction Discrepancy (Trifind vs Score > 90):"] + summary_discrepancy + [""] + [
    "📌 Sanction Reason Breakdown:"] + summary_reason

for i, line in enumerate(instructions, 1):
    ws_instr.cell(row=i, column=1).value = line

wb.save(output_path)

# === Console Output ===
print("\n📊 USAT Match Score Bin Summary:")
for _, row in summary_usat.iterrows():
    print(f"{row['match_score_bin_usat']}: {row['count']} matches")
print(f"TOTAL: {total_matches} matches")

print("\n🟢 Trifind USAT Sanctioned Flag (Raw):")
print(df_matches['trifind_usat_sanctioned_flag'].value_counts().to_string(index=True, header=False))

print("\n🟢 Inferred USAT Sanctioned (Based on Flag OR Score > 90):")
print(df_matches['inferred_usat_sanctioned'].value_counts().to_string(index=True, header=False))

print("\n⚠️ Sanction Discrepancy (Trifind vs Match Score > 90):")
print(df_matches['sanction_discrepancy_flag'].value_counts().to_string(index=True, header=False))

print("\n📌 Sanction Reason Breakdown:")
print(df_matches['reason_for_sanction'].value_counts().to_string(index=True, header=False))

# === Inferred USAT Summary by State ===
summary_by_state = (
    df_matches.groupby(['trifind_state', 'inferred_usat_sanctioned'])
    .size()
    .unstack(fill_value=0)
    .rename(columns={True: 'USAT', False: 'Non-USAT'})
)

summary_by_state['Total'] = summary_by_state.sum(axis=1)
top_states = summary_by_state.sort_values(by='Total', ascending=False).head(5)

total_events = len(df_matches)
total_usat = summary_by_state['USAT'].sum()
total_non_usat = summary_by_state['Non-USAT'].sum()

print("\n📊 Summary:")
print(f"🔢 Total Events Scraped: {total_events}")
print(f"🏆 Top 5 States by Total Events:")

for state, row in top_states.iterrows():
    print(f"{state}: {row['Total']} total — 🟢 USAT: {row['USAT']}, 🔴 Non-USAT: {row['Non-USAT']}")

print("📈 Overall Breakdown:")
print(f"🟢 USAT Sanctioned: {total_usat}")
print(f"🔴 Non-USAT: {total_non_usat}")


print(f"\n✅ Match results written to: {output_path}")
